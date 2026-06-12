from __future__ import annotations

import argparse
import math
import xml.etree.ElementTree as ElementTree
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Sequence

from Analysis.result_io import loadRecords

ALGORITHM_NAMES: tuple[str, ...] = ("drl", "ga", "greedy", "pso")
DEFAULT_RESULTS_ROOT = "__RESULTS__"
DEFAULT_OUTPUT_PATH = "__RESULTS__/analysis/n(sensor)_by_methods_report.xlsx"
METRIC_NAME = "n_inner"
XML_INDENT_SPACES = 2
STAT_NAMES: tuple[str, ...] = ("minimum", "average", "maximum", "stddev", "runs")


@dataclass(frozen=True, slots=True)
class SensorStats:
    run_count: int
    seed_bands: tuple[str, ...]
    minimum: float
    average: float
    maximum: float
    stddev: float


@dataclass(frozen=True, slots=True)
class SensorCell:
    map_name: str
    algorithm: str
    stats: SensorStats | None


def listMaps(results_root: Path, algorithms: Sequence[str]) -> list[str]:
    map_names: set[str] = set()
    for algorithm in algorithms:
        algorithm_root = results_root / algorithm
        if not algorithm_root.exists():
            continue
        map_names.update(path.name for path in algorithm_root.iterdir() if path.is_dir())
    return sorted(map_names)


def listBands(map_root: Path) -> tuple[str, ...]:
    if not map_root.exists():
        return ()
    seed_bands = [path.name for path in map_root.iterdir() if path.is_dir()]
    return tuple(sorted(seed_bands, key=bandKey))


def bandKey(seed_band: str) -> int:
    try:
        return int(seed_band.split("-", maxsplit=1)[0])
    except ValueError:
        return 0


def sensorCount(run: dict[str, object]) -> int | None:
    final_data = run.get("final", {})
    if not isinstance(final_data, dict):
        return None

    n_inner = final_data.get(METRIC_NAME)
    if isinstance(n_inner, (int, float)):
        return int(n_inner)

    solution = final_data.get("best_solution", [])
    if isinstance(solution, list):
        return len(solution)
    return None


def loadCounts(map_root: Path) -> list[int]:
    records = loadRecords(map_root)
    counts: list[int] = []
    for _, run in records:
        count = sensorCount(run)
        if count is not None:
            counts.append(count)
    return counts


def calcStats(counts: Sequence[int], seed_bands: Sequence[str]) -> SensorStats:
    if not counts:
        raise ValueError("counts must not be empty.")

    average = sum(counts) / len(counts)
    variance = sum((count - average) ** 2 for count in counts) / len(counts)
    return SensorStats(
        run_count=len(counts),
        seed_bands=tuple(seed_bands),
        minimum=float(min(counts)),
        average=float(average),
        maximum=float(max(counts)),
        stddev=float(math.sqrt(variance)),
    )


def collectCells(results_root: Path, algorithms: Sequence[str]) -> list[SensorCell]:
    cells: list[SensorCell] = []
    for map_name in listMaps(results_root, algorithms):
        for algorithm in algorithms:
            map_root = results_root / algorithm / map_name
            stats = None
            if map_root.exists():
                counts = loadCounts(map_root)
                stats = calcStats(counts, listBands(map_root))
            cells.append(SensorCell(map_name=map_name, algorithm=algorithm, stats=stats))
    return cells


def cellMap(cells: Iterable[SensorCell]) -> dict[tuple[str, str], SensorStats | None]:
    return {(cell.map_name, cell.algorithm): cell.stats for cell in cells}


def formatNumber(value: float) -> str:
    if value.is_integer():
        return str(int(value))
    return f"{value:.4f}"


def addStat(parent: ElementTree.Element, name: str, value: float) -> None:
    element = ElementTree.SubElement(parent, name)
    element.text = formatNumber(value)


def addAlgorithm(parent: ElementTree.Element, algorithm: str, stats: SensorStats | None) -> None:
    element = ElementTree.SubElement(parent, "algorithm", name=algorithm)
    if stats is None:
        element.set("status", "missing")
        return

    element.set("status", "ok")
    element.set("runs", str(stats.run_count))
    element.set("seed_bands", ",".join(stats.seed_bands))
    addStat(element, "minimum", stats.minimum)
    addStat(element, "average", stats.average)
    addStat(element, "maximum", stats.maximum)
    addStat(element, "stddev", stats.stddev)


def buildXml(results_root: Path, algorithms: Sequence[str]) -> ElementTree.Element:
    root = ElementTree.Element(
        "sensor_count_summary",
        {
            "results_root": str(results_root),
            "metric": METRIC_NAME,
            "stddev": "population",
            "generated_at_utc": datetime.now(timezone.utc).isoformat(),
        },
    )
    table = ElementTree.SubElement(root, "table", x_axis="algorithm", y_axis="map")

    cells = cellMap(collectCells(results_root, algorithms))
    for map_name in listMaps(results_root, algorithms):
        map_element = ElementTree.SubElement(table, "map", name=map_name)
        for algorithm in algorithms:
            addAlgorithm(map_element, algorithm, cells.get((map_name, algorithm)))
    return root


def saveXml(root: ElementTree.Element, output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    ElementTree.indent(root, space=" " * XML_INDENT_SPACES)
    tree = ElementTree.ElementTree(root)
    tree.write(output_path, encoding="utf-8", xml_declaration=True)


def statValue(stats: SensorStats, stat_name: str) -> float | int:
    if stat_name == "minimum":
        return stats.minimum
    if stat_name == "average":
        return stats.average
    if stat_name == "maximum":
        return stats.maximum
    if stat_name == "stddev":
        return stats.stddev
    if stat_name == "runs":
        return stats.run_count
    raise ValueError(f"unknown stat: {stat_name}")


def saveXlsx(
    *,
    cells: Sequence[SensorCell],
    algorithms: Sequence[str],
    output_path: Path,
    results_root: Path,
) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    summary = workbook.active
    summary.title = "summary"
    summary.freeze_panes = "B3"

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    subheader_fill = PatternFill("solid", fgColor="EAF3F8")
    title_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")

    summary["A1"] = "map"
    summary["A1"].font = title_font
    summary["A1"].fill = header_fill
    summary["A1"].alignment = center
    summary.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)

    column = 2
    for algorithm in algorithms:
        start_column = column
        end_column = column + len(STAT_NAMES) - 1
        summary.merge_cells(start_row=1, start_column=start_column, end_row=1, end_column=end_column)
        cell = summary.cell(row=1, column=start_column, value=algorithm)
        cell.font = title_font
        cell.fill = header_fill
        cell.alignment = center

        for stat_name in STAT_NAMES:
            stat_cell = summary.cell(row=2, column=column, value=stat_name)
            stat_cell.font = title_font
            stat_cell.fill = subheader_fill
            stat_cell.alignment = center
            column += 1

    cells_by_key = cellMap(cells)
    map_names = sorted({cell.map_name for cell in cells})
    for row_index, map_name in enumerate(map_names, start=3):
        summary.cell(row=row_index, column=1, value=map_name)
        column = 2
        for algorithm in algorithms:
            stats = cells_by_key.get((map_name, algorithm))
            for stat_name in STAT_NAMES:
                value = "" if stats is None else statValue(stats, stat_name)
                data_cell = summary.cell(row=row_index, column=column, value=value)
                if stat_name in {"average", "stddev"}:
                    data_cell.number_format = "0.0000"
                column += 1

    for column_index in range(1, summary.max_column + 1):
        letter = get_column_letter(column_index)
        summary.column_dimensions[letter].width = 14 if column_index > 1 else 18

    meta = workbook.create_sheet("metadata")
    meta.append(["key", "value"])
    meta.append(["results_root", str(results_root)])
    meta.append(["metric", METRIC_NAME])
    meta.append(["stddev", "population"])
    meta.append(["generated_at_utc", datetime.now(timezone.utc).isoformat()])
    meta.append(["seed_policy", "all seed-band directories are combined"])
    meta.append(["excluded", "__RESULTS__/combinatorial"])
    meta.column_dimensions["A"].width = 22
    meta.column_dimensions["B"].width = 70

    workbook.save(output_path)


def parseArgs() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Create a spreadsheet table of final inner sensor count statistics by map and algorithm."
    )
    parser.add_argument("--results-root", default=DEFAULT_RESULTS_ROOT)
    parser.add_argument("--output", default=DEFAULT_OUTPUT_PATH)
    parser.add_argument("--algorithms", nargs="+", default=list(ALGORITHM_NAMES))
    parser.add_argument("--format", choices=("xlsx", "xml"), default="xlsx")
    return parser.parse_args()


def main() -> None:
    args = parseArgs()
    results_root = Path(args.results_root)
    output_path = Path(args.output)
    algorithms = tuple(args.algorithms)
    if args.format == "xml":
        root = buildXml(results_root, algorithms)
        saveXml(root, output_path)
    else:
        cells = collectCells(results_root, algorithms)
        saveXlsx(cells=cells, algorithms=algorithms, output_path=output_path, results_root=results_root)
    print(output_path)


if __name__ == "__main__":
    main()
