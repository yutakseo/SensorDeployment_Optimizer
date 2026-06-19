from __future__ import annotations

import argparse
import math
import sys
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Sequence

REPO_ROOT = Path(__file__).resolve().parents[3]

if __package__ in {None, ""}:
    sys.path.insert(0, str(REPO_ROOT))

from Analysis.internal.result_io import loadRecords
from Analysis.internal.map_names import displayMapName, sortMapNames

DEFAULT_RESULTS_ROOT = "__RESULTS__"
DEFAULT_ALGORITHM = "ga"
DEFAULT_OUTPUT_PATH = "__RESULTS__/analysis/ga_result_statistics.xlsx"
METRIC_KEYS: tuple[str, ...] = ("sensor", "coverage", "runtime")
STAT_KEYS: tuple[str, ...] = ("minimum", "average", "maximum", "stddev")
METRIC_LABELS: dict[str, str] = {
    "sensor": "Sensor Count",
    "coverage": "Coverage (%)",
    "runtime": "Runtime (sec)",
}
STAT_LABELS: dict[str, str] = {
    "minimum": "Min",
    "average": "Avg",
    "maximum": "Max",
    "stddev": "StdDev",
}


@dataclass(frozen=True, slots=True)
class MetricStats:
    minimum: float
    average: float
    maximum: float
    stddev: float


@dataclass(frozen=True, slots=True)
class MapStats:
    map_name: str
    run_count: int
    seed_bands: tuple[str, ...]
    sensor: MetricStats
    coverage: MetricStats
    runtime: MetricStats


def listMaps(algorithm_root: Path) -> list[str]:
    if not algorithm_root.exists():
        return []
    return sortMapNames({path.name for path in algorithm_root.iterdir() if path.is_dir()})


def listBands(map_root: Path) -> tuple[str, ...]:
    if not map_root.exists():
        return ()
    return tuple(sorted((path.name for path in map_root.iterdir() if path.is_dir()), key=bandKey))


def bandKey(seed_band: str) -> int:
    try:
        return int(seed_band.split("-", maxsplit=1)[0])
    except ValueError:
        return 0


def numericValue(value: object) -> float | None:
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)) and math.isfinite(float(value)):
        return float(value)
    return None


def finalData(run: dict[str, object]) -> dict[str, object]:
    final = run.get("final", {})
    if isinstance(final, dict):
        return final
    return {}


def sensorCount(run: dict[str, object]) -> float | None:
    final = finalData(run)
    n_total = numericValue(final.get("n_total"))
    if n_total is not None:
        return n_total

    n_inner = numericValue(final.get("n_inner"))
    n_corner = numericValue(final.get("n_corner"))
    if n_inner is not None and n_corner is not None:
        return n_inner + n_corner

    best_solution = final.get("best_solution", [])
    corner_points = final.get("corner_points", [])
    best_count = len(best_solution) if isinstance(best_solution, list) else 0
    corner_count = len(corner_points) if isinstance(corner_points, list) else 0
    return float(best_count + corner_count)


def coverageValue(run: dict[str, object]) -> float | None:
    final = finalData(run)
    coverage = numericValue(final.get("coverage"))
    if coverage is not None:
        return coverage

    generations = run.get("generations", [])
    if not isinstance(generations, list) or not generations:
        return None

    last_generation = generations[-1]
    if not isinstance(last_generation, dict):
        return None
    return numericValue(last_generation.get("best_coverage"))


def runtimeValue(run: dict[str, object]) -> float | None:
    final = finalData(run)
    elapsed = numericValue(final.get("elapsed_sec"))
    if elapsed is not None:
        return elapsed
    return numericValue(run.get("elapsed_sec"))


def calcStats(values: Sequence[float]) -> MetricStats:
    if not values:
        raise ValueError("values must not be empty.")

    average = sum(values) / len(values)
    variance = sum((value - average) ** 2 for value in values) / len(values)
    return MetricStats(
        minimum=float(min(values)),
        average=float(average),
        maximum=float(max(values)),
        stddev=float(math.sqrt(variance)),
    )


def collectMapStats(map_root: Path, map_name: str) -> MapStats:
    records = loadRecords(map_root)
    sensor_values: list[float] = []
    coverage_values: list[float] = []
    runtime_values: list[float] = []

    for _, run in records:
        sensor_count = sensorCount(run)
        coverage = coverageValue(run)
        runtime = runtimeValue(run)

        if sensor_count is not None:
            sensor_values.append(sensor_count)
        if coverage is not None:
            coverage_values.append(coverage)
        if runtime is not None:
            runtime_values.append(runtime)

    if not sensor_values:
        raise ValueError(f"no valid sensor counts under: {map_root}")
    if not coverage_values:
        raise ValueError(f"no valid coverage values under: {map_root}")
    if not runtime_values:
        raise ValueError(f"no valid runtime values under: {map_root}")

    return MapStats(
        map_name=map_name,
        run_count=len(records),
        seed_bands=listBands(map_root),
        sensor=calcStats(sensor_values),
        coverage=calcStats(coverage_values),
        runtime=calcStats(runtime_values),
    )


def collectStats(
    *,
    results_root: Path,
    algorithm: str = DEFAULT_ALGORITHM,
) -> list[MapStats]:
    algorithm_root = results_root / algorithm
    stats_list: list[MapStats] = []
    for map_name in listMaps(algorithm_root):
        map_root = algorithm_root / map_name
        stats_list.append(collectMapStats(map_root, map_name))
    return stats_list


def resolveRepoPath(path: str | Path) -> Path:
    raw_path = Path(path)
    if raw_path.is_absolute() or raw_path.exists():
        return raw_path

    repo_path = REPO_ROOT / raw_path
    if repo_path.exists() or raw_path.parts[:1] == ("__RESULTS__",):
        return repo_path
    return raw_path


def metricStats(map_stats: MapStats, metric_key: str) -> MetricStats:
    if metric_key == "sensor":
        return map_stats.sensor
    if metric_key == "coverage":
        return map_stats.coverage
    if metric_key == "runtime":
        return map_stats.runtime
    raise ValueError(f"unknown metric: {metric_key}")


def statValue(stats: MetricStats, stat_key: str) -> float:
    if stat_key == "minimum":
        return stats.minimum
    if stat_key == "average":
        return stats.average
    if stat_key == "maximum":
        return stats.maximum
    if stat_key == "stddev":
        return stats.stddev
    raise ValueError(f"unknown stat: {stat_key}")


def addHeaders(summary) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    subheader_fill = PatternFill("solid", fgColor="EAF3F8")
    title_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")

    base_headers = ("Map", "Runs", "Seed Bands")
    for column, label in enumerate(base_headers, start=1):
        summary.merge_cells(start_row=1, start_column=column, end_row=2, end_column=column)
        cell = summary.cell(row=1, column=column, value=label)
        cell.font = title_font
        cell.fill = header_fill
        cell.alignment = center

    column = len(base_headers) + 1
    for metric_key in METRIC_KEYS:
        start_column = column
        end_column = column + len(STAT_KEYS) - 1
        summary.merge_cells(start_row=1, start_column=start_column, end_row=1, end_column=end_column)
        cell = summary.cell(row=1, column=start_column, value=METRIC_LABELS[metric_key])
        cell.font = title_font
        cell.fill = header_fill
        cell.alignment = center

        for stat_key in STAT_KEYS:
            stat_cell = summary.cell(row=2, column=column, value=STAT_LABELS[stat_key])
            stat_cell.font = title_font
            stat_cell.fill = subheader_fill
            stat_cell.alignment = center
            column += 1


def addRows(summary, stats_list: Sequence[MapStats]) -> None:
    for row_index, map_stats in enumerate(stats_list, start=3):
        summary.cell(row=row_index, column=1, value=displayMapName(map_stats.map_name))
        summary.cell(row=row_index, column=2, value=map_stats.run_count)
        summary.cell(row=row_index, column=3, value=", ".join(map_stats.seed_bands))

        column = 4
        for metric_key in METRIC_KEYS:
            stats = metricStats(map_stats, metric_key)
            for stat_key in STAT_KEYS:
                cell = summary.cell(row=row_index, column=column, value=statValue(stats, stat_key))
                if metric_key == "sensor" and stat_key in {"minimum", "maximum"}:
                    cell.number_format = "0"
                elif metric_key == "runtime":
                    cell.number_format = "0.000"
                else:
                    cell.number_format = "0.0000"
                column += 1


def formatSheet(summary) -> None:
    from openpyxl.styles import Alignment
    from openpyxl.utils import get_column_letter

    summary.freeze_panes = "D3"
    for row in summary.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")

    widths = {
        1: 18,
        2: 10,
        3: 18,
    }
    for column_index in range(1, summary.max_column + 1):
        letter = get_column_letter(column_index)
        summary.column_dimensions[letter].width = widths.get(column_index, 13)


def addMetadata(
    *,
    workbook,
    results_root: Path,
    algorithm: str,
    output_path: Path,
) -> None:
    meta = workbook.create_sheet("metadata")
    meta.append(["key", "value"])
    meta.append(["results_root", str(results_root)])
    meta.append(["algorithm", algorithm])
    meta.append(["output_path", str(output_path)])
    meta.append(["sensor_metric", "final.n_total"])
    meta.append(["coverage_metric", "final.coverage"])
    meta.append(["runtime_metric", "final.elapsed_sec"])
    meta.append(["stddev", "population"])
    meta.append(["seed_policy", "all seed-band directories are combined per map"])
    meta.append(["generated_at_utc", datetime.now(timezone.utc).isoformat()])
    meta.column_dimensions["A"].width = 22
    meta.column_dimensions["B"].width = 80


def saveXlsx(
    *,
    stats_list: Sequence[MapStats],
    output_path: Path,
    results_root: Path,
    algorithm: str = DEFAULT_ALGORITHM,
) -> None:
    from openpyxl import Workbook

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    summary = workbook.active
    summary.title = "summary"

    addHeaders(summary)
    addRows(summary, stats_list)
    formatSheet(summary)
    addMetadata(
        workbook=workbook,
        results_root=results_root,
        algorithm=algorithm,
        output_path=output_path,
    )
    workbook.save(output_path)


def saveGaStatisticsReport(
    *,
    results_root: str | Path = DEFAULT_RESULTS_ROOT,
    output_path: str | Path = DEFAULT_OUTPUT_PATH,
    algorithm: str = DEFAULT_ALGORITHM,
) -> Path:
    root_path = resolveRepoPath(results_root)
    report_path = resolveRepoPath(output_path)
    stats_list = collectStats(results_root=root_path, algorithm=algorithm)
    if not stats_list:
        raise FileNotFoundError(f"no {algorithm} map result directories under: {root_path / algorithm}")
    saveXlsx(
        stats_list=stats_list,
        output_path=report_path,
        results_root=root_path,
        algorithm=algorithm,
    )
    return report_path


def parseArgs() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Create a GA result statistics spreadsheet by map."
    )
    parser.add_argument("--results-root", default=DEFAULT_RESULTS_ROOT)
    parser.add_argument("--output", default=DEFAULT_OUTPUT_PATH)
    parser.add_argument("--algorithm", default=DEFAULT_ALGORITHM)
    return parser.parse_args()


def main() -> None:
    args = parseArgs()
    output_path = saveGaStatisticsReport(
        results_root=args.results_root,
        output_path=args.output,
        algorithm=args.algorithm,
    )
    print(output_path)


if __name__ == "__main__":
    main()
