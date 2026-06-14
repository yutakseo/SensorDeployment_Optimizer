from __future__ import annotations

import argparse
import importlib.util
import sys
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from types import ModuleType
from typing import Sequence

import numpy as np

if __package__ in {None, ""}:
    sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

DEFAULT_MAPS_ROOT = "__MAPS__"
DEFAULT_OUTPUT_PATH = "__RESULTS__/analysis/map_statistics_report.xlsx"
EXCLUDED_ROOTS: tuple[str, ...] = ("v1", "v2")
MAP_VARIABLES: tuple[str, ...] = ("MAP", "GRID")
SITE_VALUES: tuple[int, ...] = (1, 2, 3)
INSTALLABLE_VALUE = 2
UNINSTALLABLE_VALUE = 3
TARGET_VALUES: tuple[int, ...] = (2, 3)
PERCENT_MULTIPLIER = 100.0
CELL_SIZE_M = 5
CELL_AREA_M2 = CELL_SIZE_M * CELL_SIZE_M


@dataclass(frozen=True, slots=True)
class MapStats:
    map_name: str
    source_path: str
    variable_name: str
    height: int
    width: int
    total_cells: int
    site_cells: int
    site_area_m2: int
    installable_cells: int
    installable_area_m2: int
    uninstallable_cells: int
    uninstallable_area_m2: int
    target_cells: int
    target_area_m2: int
    installable_percent_of_target: float
    uninstallable_percent_of_target: float


def listMapFiles(maps_root: Path) -> list[Path]:
    if not maps_root.exists():
        raise FileNotFoundError(f"maps root not found: {maps_root}")
    return sorted(path for path in maps_root.rglob("*.py") if not shouldSkip(path, maps_root))


def shouldSkip(path: Path, maps_root: Path) -> bool:
    relative_parts = path.relative_to(maps_root).parts
    return (
        path.name == "__init__.py"
        or "__pycache__" in path.parts
        or isExcludedRoot(relative_parts)
    )


def isExcludedRoot(relative_parts: Sequence[str]) -> bool:
    if not relative_parts:
        return False
    return relative_parts[0] in EXCLUDED_ROOTS


def loadModule(path: Path) -> ModuleType:
    module_name = "_map_statistics_" + "_".join(path.with_suffix("").parts[-6:])
    spec = importlib.util.spec_from_file_location(module_name, path)
    if spec is None or spec.loader is None:
        raise ImportError(f"failed to create import spec for map file: {path}")

    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def getGrid(module: ModuleType) -> tuple[str, object] | None:
    for variable_name in MAP_VARIABLES:
        if hasattr(module, variable_name):
            return variable_name, getattr(module, variable_name)
    return None


def toMapArray(grid: object, source_path: Path) -> np.ndarray:
    try:
        map_array = np.asarray(grid)
    except ValueError as error:
        raise ValueError(f"map data must be rectangular: {source_path}") from error

    if map_array.ndim != 2:
        raise ValueError(f"map data must be 2D: {source_path}, shape={map_array.shape}")
    if not np.issubdtype(map_array.dtype, np.number):
        raise TypeError(f"map data must be numeric: {source_path}, dtype={map_array.dtype}")
    return map_array.astype(np.int64, copy=False)


def calcPercent(part: int, whole: int) -> float:
    if whole <= 0:
        return 0.0
    return part / whole * PERCENT_MULTIPLIER


def calcStats(*, map_name: str, source_path: Path, variable_name: str, map_array: np.ndarray) -> MapStats:
    height, width = (int(value) for value in map_array.shape)
    total_cells = int(map_array.size)
    site_cells = int(np.isin(map_array, SITE_VALUES).sum())
    installable_cells = int((map_array == INSTALLABLE_VALUE).sum())
    uninstallable_cells = int((map_array == UNINSTALLABLE_VALUE).sum())
    target_cells = int(np.isin(map_array, TARGET_VALUES).sum())

    return MapStats(
        map_name=map_name,
        source_path=str(source_path),
        variable_name=variable_name,
        height=height,
        width=width,
        total_cells=total_cells,
        site_cells=site_cells,
        site_area_m2=site_cells * CELL_AREA_M2,
        installable_cells=installable_cells,
        installable_area_m2=installable_cells * CELL_AREA_M2,
        uninstallable_cells=uninstallable_cells,
        uninstallable_area_m2=uninstallable_cells * CELL_AREA_M2,
        target_cells=target_cells,
        target_area_m2=target_cells * CELL_AREA_M2,
        installable_percent_of_target=calcPercent(installable_cells, target_cells),
        uninstallable_percent_of_target=calcPercent(uninstallable_cells, target_cells),
    )


def mapName(path: Path, maps_root: Path) -> str:
    relative_path = path.relative_to(maps_root).with_suffix("")
    return ".".join(relative_path.parts)


def collectStats(maps_root: Path) -> list[MapStats]:
    rows: list[MapStats] = []
    for path in listMapFiles(maps_root):
        module = loadModule(path)
        grid = getGrid(module)
        if grid is None:
            continue

        variable_name, grid_data = grid
        map_array = toMapArray(grid_data, path)
        rows.append(
            calcStats(
                map_name=mapName(path, maps_root),
                source_path=path,
                variable_name=variable_name,
                map_array=map_array,
            )
        )
    return rows


def saveXlsx(*, rows: Sequence[MapStats], output_path: Path, maps_root: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    summary = workbook.active
    summary.title = "summary"
    summary.freeze_panes = "A2"

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    title_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")

    headers = [
        "Study Area",
        "Map Height (cells)",
        "Map Width (cells)",
        "Grid Cells (cells)",
        "Site Cells (cells)",
        "Site Area (m2)",
        "Target Area (m2)",
        "Installable Area (m2)",
        "Installable Ratio (%)",
        "Restricted Area (m2)",
        "Restricted Ratio (%)",
    ]

    for column, header in enumerate(headers, start=1):
        cell = summary.cell(row=1, column=column, value=header)
        cell.font = title_font
        cell.fill = header_fill
        cell.alignment = center

    for row_index, row in enumerate(rows, start=2):
        values = [
            row.map_name,
            row.height,
            row.width,
            row.total_cells,
            row.site_cells,
            row.site_area_m2,
            row.target_area_m2,
            row.installable_area_m2,
            row.installable_percent_of_target,
            row.uninstallable_area_m2,
            row.uninstallable_percent_of_target,
        ]
        for column, value in enumerate(values, start=1):
            cell_value = round(value, 2) if isinstance(value, float) else value
            cell = summary.cell(row=row_index, column=column, value=cell_value)
            if isinstance(value, float):
                cell.number_format = "0.00"

    for column_index in range(1, summary.max_column + 1):
        letter = get_column_letter(column_index)
        summary.column_dimensions[letter].width = 18
    summary.column_dimensions["A"].width = 34
    summary.column_dimensions["E"].width = 18
    summary.column_dimensions["F"].width = 24
    summary.column_dimensions["G"].width = 24
    summary.column_dimensions["H"].width = 26
    summary.column_dimensions["I"].width = 38
    summary.column_dimensions["J"].width = 22
    summary.column_dimensions["K"].width = 38

    writeMetadata(workbook=workbook, maps_root=maps_root, row_count=len(rows))
    workbook.save(output_path)


def writeMetadata(*, workbook, maps_root: Path, row_count: int) -> None:
    meta = workbook.create_sheet("metadata")
    meta.append(["key", "value"])
    meta.append(["maps_root", str(maps_root)])
    meta.append(["excluded_roots", ",".join(EXCLUDED_ROOTS)])
    meta.append(["map_variables", ",".join(MAP_VARIABLES)])
    meta.append(["cell_size_m", CELL_SIZE_M])
    meta.append(["cell_area_m2", CELL_AREA_M2])
    meta.append(["construction_site_area_formula", "construction_site_count * 25"])
    meta.append(["target_area_formula", "target_area_count(value=2,3) * 25"])
    meta.append(["installable_area_formula", "installable_count(value=2) * 25"])
    meta.append(["uninstallable_area_formula", "uninstallable_count(value=3) * 25"])
    meta.append(["construction_site_values", ",".join(str(value) for value in SITE_VALUES)])
    meta.append(["installable_value", INSTALLABLE_VALUE])
    meta.append(["uninstallable_value", UNINSTALLABLE_VALUE])
    meta.append(["target_values", ",".join(str(value) for value in TARGET_VALUES)])
    meta.append(["percentage_basis", "installable/uninstallable percentages use target area(value=2,3) as denominator"])
    meta.append(["formula", "ratio = area / target_area_m2 * 100"])
    meta.append(["map_count", row_count])
    meta.append(["generated_at_utc", datetime.now(timezone.utc).isoformat()])
    meta.column_dimensions["A"].width = 28
    meta.column_dimensions["B"].width = 86


def parseArgs() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Create an Excel report for map dataset statistics.")
    parser.add_argument("--maps-root", default=DEFAULT_MAPS_ROOT)
    parser.add_argument("--output", default=DEFAULT_OUTPUT_PATH)
    return parser.parse_args()


def main() -> None:
    args = parseArgs()
    maps_root = Path(args.maps_root)
    output_path = Path(args.output)
    rows = collectStats(maps_root)
    saveXlsx(rows=rows, output_path=output_path, maps_root=maps_root)
    print(output_path)


if __name__ == "__main__":
    main()
