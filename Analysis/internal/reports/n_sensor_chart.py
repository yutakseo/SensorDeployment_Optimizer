from __future__ import annotations

import argparse
import math
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Sequence

if __package__ in {None, ""}:
    sys.path.insert(0, str(Path(__file__).resolve().parents[3]))

import matplotlib.pyplot as plt
import numpy as np
from matplotlib.axes import Axes
from matplotlib.container import BarContainer
from matplotlib.ticker import MaxNLocator
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from Analysis.internal.plot_style import DEFAULT_DPI, SPINE_WIDTH, applySciStyle, styleAxis
from Analysis.internal.map_names import sortMapNames
from Analysis.internal.reports.n_sensor import ALGORITHM_NAMES

DEFAULT_INPUT_PATH = "__RESULTS__/analysis/n(sensor)_by_methods_report.xlsx"
DEFAULT_OUTPUT_DIR = "__RESULTS__/analysis/n_sensor_by_algorithm"
SUMMARY_SHEET = "summary"
AVG_STAT = "Avg"
BAR_WIDTH = 0.56
FIGURE_SIZE = (3.5, 2.6)
Y_AXIS_PADDING = 1.18
MIN_Y_AXIS_TOP = 5.0
Y_AXIS_STEP = 1
BAR_COLOR = "0.70"
BAR_EDGE_COLOR = "0.15"
HIGHLIGHT_ALGORITHM = "ga"
HIGHLIGHT_COLOR = "#8EC7E8"
BAR_LABEL_SIZE = 7
LABEL_PADDING = 3


@dataclass(frozen=True, slots=True)
class SensorPoint:
    map_name: str
    algorithm: str
    average: float


@dataclass(frozen=True, slots=True)
class HeaderCell:
    algorithm: str
    stat: str


def cleanText(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def cleanNumber(value: object) -> float | None:
    if not isinstance(value, (int, float)):
        return None
    if math.isnan(float(value)):
        return None
    return float(value)


def loadHeaders(sheet: Worksheet) -> dict[int, HeaderCell]:
    headers: dict[int, HeaderCell] = {}
    algorithm = ""

    for column in range(2, sheet.max_column + 1):
        algorithm_value = cleanText(sheet.cell(row=1, column=column).value)
        stat = cleanText(sheet.cell(row=2, column=column).value)

        if algorithm_value:
            algorithm = algorithm_value

        headers[column] = HeaderCell(algorithm=algorithm, stat=stat)

    return headers


def loadSensorPoints(
    *,
    input_path: Path,
    algorithms: Sequence[str],
) -> list[SensorPoint]:
    workbook = load_workbook(input_path, data_only=True)
    if SUMMARY_SHEET not in workbook.sheetnames:
        raise ValueError(f"Workbook has no '{SUMMARY_SHEET}' sheet: {input_path}")

    sheet = workbook[SUMMARY_SHEET]
    headers = loadHeaders(sheet)
    points: list[SensorPoint] = []

    for row in range(3, sheet.max_row + 1):
        map_name = cleanText(sheet.cell(row=row, column=1).value)
        if not map_name:
            continue

        values = loadRowValues(sheet=sheet, row=row, headers=headers)
        for algorithm in algorithms:
            average = values.get((algorithm, AVG_STAT))
            if average is None:
                continue
            points.append(
                SensorPoint(
                    map_name=map_name,
                    algorithm=algorithm,
                    average=average,
                )
            )

    return points


def loadRowValues(
    sheet: Worksheet,
    row: int,
    headers: dict[int, HeaderCell],
) -> dict[tuple[str, str], float]:
    values: dict[tuple[str, str], float] = {}
    for column, header in headers.items():
        if header.stat != AVG_STAT:
            continue

        value = cleanNumber(sheet.cell(row=row, column=column).value)
        if value is None:
            continue
        values[(header.algorithm, header.stat)] = value

    return values


def groupByMap(points: Sequence[SensorPoint]) -> dict[str, list[SensorPoint]]:
    grouped: dict[str, list[SensorPoint]] = {}
    for point in points:
        grouped.setdefault(point.map_name, []).append(point)
    return grouped


def pointMap(points: Sequence[SensorPoint]) -> dict[str, SensorPoint]:
    return {point.algorithm: point for point in points}


def annotateBars(axis: Axes, bars: BarContainer, values: Sequence[float]) -> None:
    for bar, value in zip(bars, values):
        axis.annotate(
            f"{value:.1f}",
            xy=(bar.get_x() + bar.get_width() / 2, float(value)),
            xytext=(0, LABEL_PADDING),
            textcoords="offset points",
            ha="center",
            va="bottom",
            fontsize=BAR_LABEL_SIZE,
            color="black",
            clip_on=False,
            zorder=6,
            bbox={
                "boxstyle": "square,pad=0.08",
                "facecolor": "white",
                "edgecolor": "none",
                "alpha": 0.85,
            },
        )


def barColors(algorithms: Sequence[str]) -> list[str]:
    return [
        HIGHLIGHT_COLOR if algorithm == HIGHLIGHT_ALGORITHM else BAR_COLOR
        for algorithm in algorithms
    ]


def drawMap(
    axis: Axes,
    map_name: str,
    points: Sequence[SensorPoint],
    algorithms: Sequence[str],
    y_axis_top: float,
) -> None:
    by_algorithm = pointMap(points)
    available = [algorithm for algorithm in algorithms if algorithm in by_algorithm]
    x_values = np.arange(len(available))
    averages = [by_algorithm[algorithm].average for algorithm in available]

    bars = axis.bar(
        x_values,
        averages,
        BAR_WIDTH,
        color=barColors(available),
        edgecolor=BAR_EDGE_COLOR,
        linewidth=SPINE_WIDTH,
        label="Average",
        zorder=3,
    )
    annotateBars(axis, bars, averages)

    axis.set_ylabel("Number of sensors")
    axis.set_xticks(x_values)
    axis.set_xticklabels([algorithm.upper() for algorithm in available])
    axis.set_ylim(0, y_axis_top)
    axis.yaxis.set_major_locator(MaxNLocator(integer=True))
    axis.set_axisbelow(True)
    axis.grid(True, axis="y", color="0.88", linewidth=0.5)
    styleAxis(axis)


def chartPath(output_dir: Path, map_name: str) -> Path:
    return output_dir / f"{map_name}.png"


def saveMapChart(
    *,
    output_dir: Path,
    map_name: str,
    points: Sequence[SensorPoint],
    algorithms: Sequence[str],
    y_axis_top: float,
    dpi: int,
    show: bool,
) -> Path:
    applySciStyle()
    fig, axis = plt.subplots(figsize=FIGURE_SIZE, dpi=dpi)
    drawMap(axis, map_name, points, algorithms, y_axis_top)
    fig.tight_layout()

    output_path = chartPath(output_dir, map_name)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(output_path, dpi=dpi)
    if show:
        plt.show()
    plt.close(fig)
    return output_path


def calcYAxisTop(points: Sequence[SensorPoint]) -> float:
    max_average = max(point.average for point in points)
    padded_top = max(MIN_Y_AXIS_TOP, max_average * Y_AXIS_PADDING)
    return float(math.ceil(padded_top / Y_AXIS_STEP) * Y_AXIS_STEP)


def saveSensorChart(
    *,
    input_path: str = DEFAULT_INPUT_PATH,
    output_dir: str = DEFAULT_OUTPUT_DIR,
    algorithms: Sequence[str] = ALGORITHM_NAMES,
    dpi: int = DEFAULT_DPI,
    show: bool = False,
) -> list[Path]:
    """Create one per-map sensor-count chart by algorithm."""
    source = Path(input_path)
    target_dir = Path(output_dir)
    points = loadSensorPoints(input_path=source, algorithms=algorithms)
    if not points:
        raise ValueError(f"No plottable sensor-count rows found: {source}")

    grouped = groupByMap(points)
    y_axis_top = calcYAxisTop(points)
    target_dir.mkdir(parents=True, exist_ok=True)
    return [
        saveMapChart(
            output_dir=target_dir,
            map_name=map_name,
            points=grouped[map_name],
            algorithms=algorithms,
            y_axis_top=y_axis_top,
            dpi=dpi,
            show=show,
        )
        for map_name in sortMapNames(set(grouped))
    ]


def parseArgs() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Create per-map sensor-count charts from the sensor count Excel report."
    )
    parser.add_argument("--input", default=DEFAULT_INPUT_PATH)
    parser.add_argument("--output-dir", default=DEFAULT_OUTPUT_DIR)
    parser.add_argument("--algorithms", nargs="+", default=list(ALGORITHM_NAMES))
    parser.add_argument("--dpi", type=int, default=DEFAULT_DPI)
    parser.add_argument("--show", action="store_true")
    return parser.parse_args()


def main() -> None:
    args = parseArgs()
    output_paths = saveSensorChart(
        input_path=args.input,
        output_dir=args.output_dir,
        algorithms=tuple(args.algorithms),
        dpi=args.dpi,
        show=args.show,
    )
    for output_path in output_paths:
        print(output_path)


if __name__ == "__main__":
    main()
