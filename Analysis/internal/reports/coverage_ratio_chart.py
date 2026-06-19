from __future__ import annotations

import argparse
import math
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Sequence

if __package__ in {None, ""}:
    sys.path.insert(0, str(Path(__file__).resolve().parents[3]))

import matplotlib.pyplot as plt
import numpy as np
from matplotlib.axes import Axes
from matplotlib.container import BarContainer
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from Analysis.internal.plot_style import DEFAULT_DPI, SPINE_WIDTH, applySciStyle, styleAxis
from Analysis.internal.map_names import sortMapNames
from Analysis.internal.reports.coverage_ratio import ALGORITHM_NAMES

DEFAULT_INPUT_PATH = "__RESULTS__/analysis/ratio(coverage)_report.xlsx"
DEFAULT_OUTPUT_DIR = "__RESULTS__/analysis/coverage_overlap_by_algorithm"
SUMMARY_SHEET = "summary"
COVERAGE_METRIC = "coverage"
OVERLAP_METRIC = "overlap"
AVERAGE_STAT = "avg"
BAR_WIDTH = 0.56
Y_AXIS_TOP = 100.0
FIGURE_SIZE = (3.5, 2.6)
LABEL_PADDING = 3
BAR_LABEL_SIZE = 7
COVERAGE_COLOR = "0.70"
OVERLAP_COLOR = "white"
BAR_EDGE_COLOR = "0.15"
OVERLAP_HATCH = "////"


@dataclass(frozen=True, slots=True)
class RatioPoint:
    map_name: str
    algorithm: str
    coverage: float
    overlap: float


@dataclass(frozen=True, slots=True)
class HeaderCell:
    algorithm: str
    metric: str
    stat: str


def cleanText(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def cleanNumber(value: object) -> float | None:
    if not isinstance(value, (int, float)):
        return None
    if math.isnan(float(value)):
        return None
    return float(value)


def loadHeaders(sheet: Worksheet) -> dict[int, HeaderCell]:
    headers: dict[int, HeaderCell] = {}
    algorithm = ""
    metric = ""

    for column in range(2, sheet.max_column + 1):
        algorithm_value = cleanText(sheet.cell(row=1, column=column).value)
        metric_value = cleanText(sheet.cell(row=2, column=column).value)
        stat = cleanText(sheet.cell(row=3, column=column).value)

        if algorithm_value:
            algorithm = algorithm_value
        if metric_value:
            metric = metric_value

        headers[column] = HeaderCell(algorithm=algorithm, metric=metric, stat=stat)

    return headers


def loadRatioPoints(
    *,
    input_path: Path,
    algorithms: Sequence[str],
) -> list[RatioPoint]:
    workbook = load_workbook(input_path, data_only=True)
    if SUMMARY_SHEET not in workbook.sheetnames:
        raise ValueError(f"Workbook has no '{SUMMARY_SHEET}' sheet: {input_path}")

    sheet = workbook[SUMMARY_SHEET]
    headers = loadHeaders(sheet)
    points: list[RatioPoint] = []

    for row in range(4, sheet.max_row + 1):
        map_name = cleanText(sheet.cell(row=row, column=1).value)
        if not map_name:
            continue

        values = loadRowValues(sheet=sheet, row=row, headers=headers)
        for algorithm in algorithms:
            coverage = values.get((algorithm, COVERAGE_METRIC))
            overlap = values.get((algorithm, OVERLAP_METRIC))
            if coverage is None or overlap is None:
                continue
            points.append(
                RatioPoint(
                    map_name=map_name,
                    algorithm=algorithm,
                    coverage=coverage,
                    overlap=overlap,
                )
            )

    return points


def loadRowValues(
    sheet: Worksheet,
    row: int,
    headers: dict[int, HeaderCell],
) -> dict[tuple[str, str], float]:
    values: dict[tuple[str, str], float] = {}
    for column, header in headers.items():
        if header.stat != AVERAGE_STAT:
            continue
        if header.metric not in {COVERAGE_METRIC, OVERLAP_METRIC}:
            continue

        value = cleanNumber(sheet.cell(row=row, column=column).value)
        if value is None:
            continue
        values[(header.algorithm, header.metric)] = value

    return values


def groupByMap(points: Sequence[RatioPoint]) -> dict[str, list[RatioPoint]]:
    grouped: dict[str, list[RatioPoint]] = {}
    for point in points:
        grouped.setdefault(point.map_name, []).append(point)
    return grouped


def pointMap(points: Sequence[RatioPoint]) -> dict[str, RatioPoint]:
    return {point.algorithm: point for point in points}


def annotateBars(
    axis: Axes,
    bars: BarContainer,
    *,
    color: str,
) -> None:
    for bar in bars:
        height = float(bar.get_height())
        axis.annotate(
            f"{height:.1f}",
            xy=(bar.get_x() + bar.get_width() / 2, height),
            xytext=(0, LABEL_PADDING),
            textcoords="offset points",
            ha="center",
            va="bottom",
            fontsize=BAR_LABEL_SIZE,
            color=color,
            clip_on=False,
        )


def drawMap(
    axis: Axes,
    map_name: str,
    points: Sequence[RatioPoint],
    algorithms: Sequence[str],
) -> None:
    by_algorithm = pointMap(points)
    available = [algorithm for algorithm in algorithms if algorithm in by_algorithm]
    x = np.arange(len(available))
    coverage = [by_algorithm[algorithm].coverage for algorithm in available]
    overlap = [by_algorithm[algorithm].overlap for algorithm in available]

    coverage_bars = axis.bar(
        x,
        coverage,
        BAR_WIDTH,
        color=COVERAGE_COLOR,
        edgecolor=BAR_EDGE_COLOR,
        linewidth=SPINE_WIDTH,
        label="Coverage (%)",
    )
    overlap_bars = axis.bar(
        x,
        overlap,
        BAR_WIDTH,
        color=OVERLAP_COLOR,
        edgecolor=BAR_EDGE_COLOR,
        hatch=OVERLAP_HATCH,
        linewidth=SPINE_WIDTH,
        label="Overlap (%)",
        zorder=3,
    )

    annotateBars(
        axis,
        coverage_bars,
        color="black",
    )
    annotateBars(
        axis,
        overlap_bars,
        color="black",
    )
    axis.set_ylabel("Percent (%)")
    axis.set_xticks(x)
    axis.set_xticklabels([algorithm.upper() for algorithm in available])
    axis.set_ylim(0, Y_AXIS_TOP)
    axis.set_axisbelow(True)
    axis.grid(True, axis="y", color="0.88", linewidth=0.5)
    axis.legend(
        frameon=False,
        loc="lower center",
        ncol=2,
        bbox_to_anchor=(0.5, 1.08),
        borderaxespad=0.0,
        handlelength=1.4,
    )
    styleAxis(axis)


def chartPath(output_dir: Path, map_name: str) -> Path:
    return output_dir / f"{map_name}.png"


def saveMapChart(
    *,
    output_dir: Path,
    map_name: str,
    points: Sequence[RatioPoint],
    algorithms: Sequence[str],
    dpi: int,
    show: bool,
) -> Path:
    applySciStyle()
    fig, axis = plt.subplots(figsize=FIGURE_SIZE, dpi=dpi)
    drawMap(axis, map_name, points, algorithms)
    fig.tight_layout()

    output_path = chartPath(output_dir, map_name)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(output_path, dpi=dpi)
    if show:
        plt.show()
    plt.close(fig)
    return output_path


def saveCoverageChart(
    *,
    input_path: str = DEFAULT_INPUT_PATH,
    output_dir: str = DEFAULT_OUTPUT_DIR,
    algorithms: Sequence[str] = ALGORITHM_NAMES,
    dpi: int = DEFAULT_DPI,
    show: bool = False,
) -> list[Path]:
    """Create one overlaid coverage/overlap bar chart per map."""
    source = Path(input_path)
    target_dir = Path(output_dir)
    points = loadRatioPoints(input_path=source, algorithms=algorithms)
    if not points:
        raise ValueError(f"No plottable coverage ratio rows found: {source}")

    grouped = groupByMap(points)
    target_dir.mkdir(parents=True, exist_ok=True)
    return [
        saveMapChart(
            output_dir=target_dir,
            map_name=map_name,
            points=grouped[map_name],
            algorithms=algorithms,
            dpi=dpi,
            show=show,
        )
        for map_name in sortMapNames(set(grouped))
    ]


def parseArgs() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Create per-map coverage/overlap bar charts from the ratio Excel report."
    )
    parser.add_argument("--input", default=DEFAULT_INPUT_PATH)
    parser.add_argument("--output-dir", default=DEFAULT_OUTPUT_DIR)
    parser.add_argument("--algorithms", nargs="+", default=list(ALGORITHM_NAMES))
    parser.add_argument("--dpi", type=int, default=DEFAULT_DPI)
    parser.add_argument("--show", action="store_true")
    return parser.parse_args()


def main() -> None:
    args = parseArgs()
    output_paths = saveCoverageChart(
        input_path=args.input,
        output_dir=args.output_dir,
        algorithms=tuple(args.algorithms),
        dpi=args.dpi,
        show=args.show,
    )
    for output_path in output_paths:
        print(output_path)


if __name__ == "__main__":
    main()
