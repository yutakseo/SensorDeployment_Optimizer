from __future__ import annotations

import argparse
import sys
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Sequence

if __package__ in {None, ""}:
    sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from Analysis.internal.result_io import bandKey, loadRecords
from Analysis.internal.trends import analyzeChange, buildTrend, getSensorSeries

ALGORITHM_NAMES: tuple[str, ...] = ("drl", "ga", "greedy", "pso")
DEFAULT_RESULTS_ROOT = "__RESULTS__"
DEFAULT_OUTPUT_PATH = "__RESULTS__/analysis/Trend_report.xlsx"
DEFAULT_THRESHOLD = 0.5
DEFAULT_METRIC = "best"
DEFAULT_INCLUDE_CORNERS = True
MISSING_VALUE = ""


@dataclass(frozen=True, slots=True)
class TrendStats:
    algorithm: str
    map_name: str
    seed_band: str
    run_count: int
    generation_count: int
    initial_mean: float
    final_mean: float
    final_std: float
    convergence_gen: int
    max_change: float
    mean_series: tuple[float, ...]
    std_series: tuple[float, ...]


def listAlgorithms(results_root: Path, algorithms: Sequence[str]) -> tuple[str, ...]:
    available = [
        algorithm
        for algorithm in algorithms
        if (results_root / algorithm).exists()
    ]
    return tuple(available or algorithms)


def listMaps(results_root: Path, algorithms: Sequence[str]) -> list[str]:
    map_names: set[str] = set()
    for algorithm in algorithms:
        algorithm_root = results_root / algorithm
        if not algorithm_root.exists():
            continue
        map_names.update(path.name for path in algorithm_root.iterdir() if path.is_dir())
    return sorted(map_names)


def listBands(map_root: Path) -> tuple[str, ...]:
    if not map_root.exists():
        return ()
    seed_bands = [path.name for path in map_root.iterdir() if path.is_dir()]
    return tuple(sorted(seed_bands, key=bandKey))


def collectBandTrend(
    *,
    results_root: Path,
    algorithm: str,
    map_name: str,
    seed_band: str,
    include_corners: bool,
    metric: str,
    threshold: float,
) -> TrendStats | None:
    map_root = results_root / algorithm / map_name
    records_root = map_root / seed_band if seed_band else map_root
    try:
        records = loadRecords(records_root)
    except (FileNotFoundError, ValueError):
        return None

    run_series = [
        getSensorSeries(run, include_corners=include_corners, metric=metric)
        for _, run in records
    ]
    mean_series, std_series = buildTrend(run_series)
    if not mean_series:
        return None

    band_label = seed_band or "all"
    trend_info = {
        band_label: {
            "runs": len(run_series),
            "gens": len(mean_series),
            "mean": mean_series,
            "std": std_series,
        }
    }
    convergence = analyzeChange(trend_info, threshold=threshold, verbose=False)
    band_convergence = convergence["by_band"].get(band_label)
    convergence_gen = 0
    max_change = 0.0
    if band_convergence is not None:
        convergence_gen = int(band_convergence["gen_from"])
        max_change = float(band_convergence["max_change"])

    return TrendStats(
        algorithm=algorithm,
        map_name=map_name,
        seed_band=band_label,
        run_count=len(run_series),
        generation_count=len(mean_series),
        initial_mean=float(mean_series[0]),
        final_mean=float(mean_series[-1]),
        final_std=float(std_series[-1]),
        convergence_gen=convergence_gen,
        max_change=max_change,
        mean_series=tuple(float(value) for value in mean_series),
        std_series=tuple(float(value) for value in std_series),
    )


def collectTrends(
    *,
    results_root: Path,
    algorithms: Sequence[str],
    include_corners: bool,
    metric: str,
    threshold: float,
) -> list[TrendStats]:
    trends: list[TrendStats] = []
    for algorithm in algorithms:
        for map_name in listMaps(results_root, [algorithm]):
            map_root = results_root / algorithm / map_name
            seed_bands = listBands(map_root)
            if not seed_bands:
                seed_bands = ("",)
            for seed_band in seed_bands:
                trend = collectBandTrend(
                    results_root=results_root,
                    algorithm=algorithm,
                    map_name=map_name,
                    seed_band=seed_band,
                    include_corners=include_corners,
                    metric=metric,
                    threshold=threshold,
                )
                if trend is not None:
                    trends.append(trend)
    return trends


def saveXlsx(
    *,
    trends: Sequence[TrendStats],
    output_path: Path,
    results_root: Path,
    include_corners: bool,
    metric: str,
    threshold: float,
) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    summary = workbook.active
    summary.title = "summary"
    writeSummary(summary, trends)
    writeTrendData(workbook.create_sheet("trend_data"), trends)
    writeMetadata(
        workbook=workbook,
        results_root=results_root,
        include_corners=include_corners,
        metric=metric,
        threshold=threshold,
    )

    for sheet in workbook.worksheets:
        styleSheet(sheet)

    workbook.save(output_path)


def writeSummary(summary, trends: Sequence[TrendStats]) -> None:
    headers = [
        "algorithm",
        "map",
        "seed_band",
        "runs",
        "generations",
        "initial_mean",
        "final_mean",
        "final_std",
        "convergence_gen",
        "max_change",
    ]
    summary.append(headers)
    for trend in trends:
        summary.append(
            [
                trend.algorithm,
                trend.map_name,
                trend.seed_band,
                trend.run_count,
                trend.generation_count,
                trend.initial_mean,
                trend.final_mean,
                trend.final_std,
                trend.convergence_gen,
                trend.max_change,
            ]
        )
    summary.freeze_panes = "A2"


def writeTrendData(sheet, trends: Sequence[TrendStats]) -> None:
    sheet.append(["algorithm", "map", "seed_band", "generation", "mean", "std"])
    for trend in trends:
        for index, mean_value in enumerate(trend.mean_series, start=1):
            std_value = trend.std_series[index - 1] if index <= len(trend.std_series) else MISSING_VALUE
            sheet.append(
                [
                    trend.algorithm,
                    trend.map_name,
                    trend.seed_band,
                    index,
                    mean_value,
                    std_value,
                ]
            )
    sheet.freeze_panes = "A2"


def writeMetadata(
    *,
    workbook,
    results_root: Path,
    include_corners: bool,
    metric: str,
    threshold: float,
) -> None:
    meta = workbook.create_sheet("metadata")
    meta.append(["key", "value"])
    meta.append(["results_root", str(results_root)])
    meta.append(["metric", metric])
    meta.append(["include_corners", str(include_corners)])
    meta.append(["threshold", float(threshold)])
    meta.append(["trend_std", "sample"])
    meta.append(["generated_at_utc", datetime.now(timezone.utc).isoformat()])
    meta.append(["seed_policy", "seed-band directories are reported separately"])


def styleSheet(sheet) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    title_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")

    for cell in sheet[1]:
        cell.font = title_font
        cell.fill = header_fill
        cell.alignment = center

    for column_index in range(1, sheet.max_column + 1):
        letter = get_column_letter(column_index)
        sheet.column_dimensions[letter].width = 16

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            if isinstance(cell.value, float):
                cell.number_format = "0.0000"


def parseArgs() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Create an Excel report for sensor-count convergence trends."
    )
    parser.add_argument("--results-root", default=DEFAULT_RESULTS_ROOT)
    parser.add_argument("--output", default=DEFAULT_OUTPUT_PATH)
    parser.add_argument("--algorithms", nargs="+", default=None)
    parser.add_argument("--metric", default=DEFAULT_METRIC)
    parser.add_argument("--threshold", type=float, default=DEFAULT_THRESHOLD)
    parser.add_argument("--inner-only", action="store_true")
    return parser.parse_args()


def main() -> None:
    args = parseArgs()
    results_root = Path(args.results_root)
    algorithms = tuple(args.algorithms) if args.algorithms else listAlgorithms(
        results_root,
        ALGORITHM_NAMES,
    )
    trends = collectTrends(
        results_root=results_root,
        algorithms=algorithms,
        include_corners=not args.inner_only,
        metric=str(args.metric),
        threshold=float(args.threshold),
    )
    output_path = Path(args.output)
    saveXlsx(
        trends=trends,
        output_path=output_path,
        results_root=results_root,
        include_corners=not args.inner_only,
        metric=str(args.metric),
        threshold=float(args.threshold),
    )
    print(output_path)


if __name__ == "__main__":
    main()
