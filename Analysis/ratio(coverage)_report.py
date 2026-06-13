from __future__ import annotations

import argparse
import math
import sys
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Sequence

if __package__ in {None, ""}:
    sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from Analysis.trends import loadOverlapRows
from Engine.map_loader import MapLoader

ALGORITHM_NAMES: tuple[str, ...] = ("drl", "ga", "greedy", "pso")
DEFAULT_RESULTS_ROOT = "__RESULTS__"
DEFAULT_OUTPUT_PATH = "__RESULTS__/analysis/ratio(coverage)_report.xlsx"
COVERAGE_KEY = "coverage_percent"
OVERLAP_KEY = "overlap_percent_of_covered"
STAT_NAMES: tuple[str, ...] = (
    "runs",
)
METRIC_NAMES: tuple[str, ...] = ("coverage", "overlap")
METRIC_STAT_NAMES: tuple[str, ...] = ("min", "avg", "max", "std")
TARGET_VALUES: tuple[int, ...] = (2, 3)


@dataclass(frozen=True, slots=True)
class MetricStats:
    minimum: float
    average: float
    maximum: float
    stddev: float


@dataclass(frozen=True, slots=True)
class ReportStats:
    run_count: int
    seed_bands: tuple[str, ...]
    coverage: MetricStats
    overlap: MetricStats


@dataclass(frozen=True, slots=True)
class ReportCell:
    map_name: str
    algorithm: str
    stats: ReportStats | None


def listMaps(results_root: Path, algorithms: Sequence[str]) -> list[str]:
    map_names: set[str] = set()
    for algorithm in algorithms:
        algorithm_root = results_root / algorithm
        if not algorithm_root.exists():
            continue
        map_names.update(path.name for path in algorithm_root.iterdir() if path.is_dir())
    return sorted(map_names)


def listAlgorithms(results_root: Path, algorithms: Sequence[str]) -> tuple[str, ...]:
    available = [
        algorithm
        for algorithm in algorithms
        if (results_root / algorithm).exists()
    ]
    return tuple(available or algorithms)


def listBands(map_root: Path) -> tuple[str, ...]:
    if not map_root.exists():
        return ()
    seed_bands = [path.name for path in map_root.iterdir() if path.is_dir()]
    return tuple(sorted(seed_bands, key=bandKey))


def bandKey(seed_band: str) -> int:
    try:
        return int(seed_band.split("-", maxsplit=1)[0])
    except ValueError:
        return 0


def cleanValues(values: Sequence[float]) -> list[float]:
    return [float(value) for value in values if not math.isnan(float(value))]


def calcMetric(values: Sequence[float]) -> MetricStats:
    clean = cleanValues(values)
    if not clean:
        return MetricStats(
            minimum=float("nan"),
            average=float("nan"),
            maximum=float("nan"),
            stddev=float("nan"),
        )

    average = sum(clean) / len(clean)
    variance = sum((value - average) ** 2 for value in clean) / len(clean)
    return MetricStats(
        minimum=min(clean),
        average=average,
        maximum=max(clean),
        stddev=math.sqrt(variance),
    )


def loadStats(results_root: Path, algorithm: str, map_name: str) -> ReportStats | None:
    map_root = results_root / algorithm / map_name
    if not map_root.exists():
        return None

    map_data = MapLoader().load(map_name)
    rows = loadOverlapRows(
        results_root=str(results_root),
        algorithm=algorithm,
        map_name=map_name,
        map_data=map_data,
        target_values=TARGET_VALUES,
        seed_band=None,
    )
    if not rows:
        return None

    coverage_values = [row[COVERAGE_KEY] for row in rows]
    overlap_values = [row[OVERLAP_KEY] for row in rows]
    return ReportStats(
        run_count=len(rows),
        seed_bands=listBands(map_root),
        coverage=calcMetric(coverage_values),
        overlap=calcMetric(overlap_values),
    )


def collectCells(results_root: Path, algorithms: Sequence[str]) -> list[ReportCell]:
    cells: list[ReportCell] = []
    for map_name in listMaps(results_root, algorithms):
        for algorithm in algorithms:
            cells.append(
                ReportCell(
                    map_name=map_name,
                    algorithm=algorithm,
                    stats=loadStats(results_root, algorithm, map_name),
                )
            )
    return cells


def cellMap(cells: Sequence[ReportCell]) -> dict[tuple[str, str], ReportStats | None]:
    return {(cell.map_name, cell.algorithm): cell.stats for cell in cells}


def metricValue(stats: ReportStats, metric_name: str, stat_name: str) -> float:
    if metric_name == "coverage" and stat_name == "min":
        return stats.coverage.minimum
    if metric_name == "coverage" and stat_name == "avg":
        return stats.coverage.average
    if metric_name == "coverage" and stat_name == "max":
        return stats.coverage.maximum
    if metric_name == "coverage" and stat_name == "std":
        return stats.coverage.stddev
    if metric_name == "overlap" and stat_name == "min":
        return stats.overlap.minimum
    if metric_name == "overlap" and stat_name == "avg":
        return stats.overlap.average
    if metric_name == "overlap" and stat_name == "max":
        return stats.overlap.maximum
    if metric_name == "overlap" and stat_name == "std":
        return stats.overlap.stddev
    raise ValueError(f"unknown metric/stat: {metric_name}/{stat_name}")


def statValue(stats: ReportStats, stat_name: str) -> int:
    if stat_name == "runs":
        return stats.run_count
    raise ValueError(f"unknown stat: {stat_name}")


def saveXlsx(
    *,
    cells: Sequence[ReportCell],
    algorithms: Sequence[str],
    output_path: Path,
    results_root: Path,
) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    summary = workbook.active
    summary.title = "summary"
    summary.freeze_panes = "B4"

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    subheader_fill = PatternFill("solid", fgColor="EAF3F8")
    title_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")

    summary["A1"] = "map"
    summary["A1"].font = title_font
    summary["A1"].fill = header_fill
    summary["A1"].alignment = center
    summary.merge_cells(start_row=1, start_column=1, end_row=3, end_column=1)

    column = 2
    for algorithm in algorithms:
        start_column = column
        metric_columns = len(METRIC_NAMES) * len(METRIC_STAT_NAMES)
        end_column = column + metric_columns + len(STAT_NAMES) - 1
        summary.merge_cells(start_row=1, start_column=start_column, end_row=1, end_column=end_column)
        cell = summary.cell(row=1, column=start_column, value=algorithm)
        cell.font = title_font
        cell.fill = header_fill
        cell.alignment = center

        for metric_name in METRIC_NAMES:
            metric_start = column
            metric_end = column + len(METRIC_STAT_NAMES) - 1
            summary.merge_cells(
                start_row=2,
                start_column=metric_start,
                end_row=2,
                end_column=metric_end,
            )
            metric_cell = summary.cell(row=2, column=metric_start, value=metric_name)
            metric_cell.font = title_font
            metric_cell.fill = subheader_fill
            metric_cell.alignment = center

            for stat_name in METRIC_STAT_NAMES:
                stat_cell = summary.cell(row=3, column=column, value=stat_name)
                stat_cell.font = title_font
                stat_cell.fill = subheader_fill
                stat_cell.alignment = center
                column += 1

        for stat_name in STAT_NAMES:
            summary.merge_cells(start_row=2, start_column=column, end_row=3, end_column=column)
            stat_cell = summary.cell(row=2, column=column, value=stat_name)
            stat_cell.font = title_font
            stat_cell.fill = subheader_fill
            stat_cell.alignment = center
            column += 1

    writeRows(summary=summary, cells=cells, algorithms=algorithms)
    setWidths(summary)
    writeMetadata(workbook=workbook, results_root=results_root)
    workbook.save(output_path)


def writeRows(summary, cells: Sequence[ReportCell], algorithms: Sequence[str]) -> None:
    cells_by_key = cellMap(cells)
    map_names = sorted({cell.map_name for cell in cells})
    for row_index, map_name in enumerate(map_names, start=4):
        summary.cell(row=row_index, column=1, value=map_name)
        column = 2
        for algorithm in algorithms:
            stats = cells_by_key.get((map_name, algorithm))
            for metric_name in METRIC_NAMES:
                for stat_name in METRIC_STAT_NAMES:
                    value = "" if stats is None else metricValue(stats, metric_name, stat_name)
                    data_cell = summary.cell(row=row_index, column=column, value=value)
                    data_cell.number_format = "0.0000"
                    column += 1
            for stat_name in STAT_NAMES:
                value = "" if stats is None else statValue(stats, stat_name)
                summary.cell(row=row_index, column=column, value=value)
                column += 1


def setWidths(summary) -> None:
    from openpyxl.utils import get_column_letter

    for column_index in range(1, summary.max_column + 1):
        letter = get_column_letter(column_index)
        summary.column_dimensions[letter].width = 15 if column_index > 1 else 18


def writeMetadata(*, workbook, results_root: Path) -> None:
    meta = workbook.create_sheet("metadata")
    meta.append(["key", "value"])
    meta.append(["results_root", str(results_root)])
    meta.append(["coverage_metric", COVERAGE_KEY])
    meta.append(["overlap_metric", OVERLAP_KEY])
    meta.append(["target_values", ",".join(str(value) for value in TARGET_VALUES)])
    meta.append(["stddev", "population"])
    meta.append(["generated_at_utc", datetime.now(timezone.utc).isoformat()])
    meta.append(["seed_policy", "all seed-band directories are combined"])
    meta.append(["excluded", "__RESULTS__/combinatorial"])
    meta.column_dimensions["A"].width = 22
    meta.column_dimensions["B"].width = 78


def parseArgs() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Create an Excel report for coverage and overlap statistics by map and algorithm."
    )
    parser.add_argument("--results-root", default=DEFAULT_RESULTS_ROOT)
    parser.add_argument("--output", default=DEFAULT_OUTPUT_PATH)
    parser.add_argument("--algorithms", nargs="+", default=None)
    return parser.parse_args()


def main() -> None:
    args = parseArgs()
    results_root = Path(args.results_root)
    output_path = Path(args.output)
    algorithms = tuple(args.algorithms) if args.algorithms else listAlgorithms(
        results_root,
        ALGORITHM_NAMES,
    )
    cells = collectCells(results_root, algorithms)
    saveXlsx(cells=cells, algorithms=algorithms, output_path=output_path, results_root=results_root)
    print(output_path)


if __name__ == "__main__":
    main()
